"""
Validation Report Generator
============================
Generates Excel and PDF validation reports from BVEngine API results.

Usage:
  python generate_report.py --result result.json --format excel
  python generate_report.py --result result.json --format pdf
  python generate_report.py --result result.json --format both

Or call directly after parse_and_validate.py:
  python parse_and_validate.py --file data.pdf --save
  python generate_report.py --result data_result.json --format both
"""

import os
import json
import argparse
from datetime import datetime
from pathlib import Path


# ── Excel Report ───────────────────────────────────────────────────────────────
def generate_excel(result: dict, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # colours
    SE_GREEN    = "3DCD58"
    DARK        = "1C1B18"
    HIGH_RED    = "F85149"
    MED_AMBER   = "D29922"
    LOW_GREEN   = "27AE60"
    LIGHT_GREY  = "F4F3EF"
    MID_GREY    = "E2E0D8"
    WHITE       = "FFFFFF"

    sev         = result.get("severity", "LOW")
    sev_colour  = {"HIGH": HIGH_RED, "MEDIUM": MED_AMBER, "LOW": LOW_GREEN}.get(sev, LOW_GREEN)

    def cell_style(ws, coord, value=None, bold=False, size=11,
                   color=DARK, bg=None, align="left", wrap=False, border=False):
        c = ws[coord]
        if value is not None:
            c.value = value
        c.font      = Font(name="Arial", bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        if border:
            thin = Side(style="thin", color=MID_GREY)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def set_col_widths(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False

    # header band
    ws1.row_dimensions[1].height = 50
    ws1.merge_cells("A1:H1")
    h = ws1["A1"]
    h.value     = "BILLING VALIDATION REPORT"
    h.font      = Font(name="Arial", bold=True, size=16, color=WHITE)
    h.fill      = PatternFill("solid", start_color=DARK)
    h.alignment = Alignment(horizontal="left", vertical="center",
                            indent=2)

    # sub header
    ws1.row_dimensions[2].height = 28
    ws1.merge_cells("A2:H2")
    sub = ws1["A2"]
    sub.value     = f"Schneider Electric · MPS · BVEngine v2.0   |   Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    sub.font      = Font(name="Arial", size=9, color="888880")
    sub.fill      = PatternFill("solid", start_color="F4F3EF")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # severity banner
    ws1.row_dimensions[3].height = 40
    ws1.merge_cells("A3:H3")
    banner = ws1["A3"]
    banner.value     = f"  {sev} SEVERITY  ·  Risk Score: {result.get('riskScore',0)}  ·  {result.get('anomalyType','None')}  ·  Recommendation: {result.get('recommendation','—')}"
    banner.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    banner.fill      = PatternFill("solid", start_color=sev_colour)
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # ── Key metrics ────────────────────────────────────────────────────────────
    ws1.row_dimensions[4].height = 8

    def info_row(ws, row, label, value, value_color=DARK):
        ws.row_dimensions[row].height = 22
        cell_style(ws, f"A{row}", label, color="888880", size=9)
        c = ws[f"B{row}"]
        c.value     = value
        c.font      = Font(name="Arial", bold=True, size=10, color=value_color)
        c.alignment = Alignment(horizontal="left", vertical="center")

    info_row(ws1, 5,  "Meter ID",         result.get("meterId","—"))
    info_row(ws1, 6,  "Billing Period",    result.get("billingPeriod","—"))
    info_row(ws1, 7,  "Processed At",      result.get("processedAt","—")[:19].replace("T"," ") + " UTC")
    info_row(ws1, 8,  "History Months",    f"{result.get('historyMonths',0)} months  ·  Model confidence: {result.get('modelConfidence','—')}")
    info_row(ws1, 9,  "Actual Energy",     f"{result.get('actualWh',0):,.0f} Wh")
    info_row(ws1, 10, "Predicted Energy",  f"{result.get('predictedWh',0):,.0f} Wh")
    dev = result.get("deviationPct", 0)
    info_row(ws1, 11, "Deviation",         f"{'+' if dev>0 else ''}{dev:.1f}%",
             HIGH_RED if dev < -5 else MED_AMBER if dev > 5 else LOW_GREEN)
    info_row(ws1, 12, "Historical Avg",    f"{result.get('historicalAvgWh',0):,.0f} Wh")
    info_row(ws1, 13, "Revenue at Risk",   f"Rs {result.get('revenueAtRiskINR',0):,.2f}",
             HIGH_RED if result.get("revenueAtRiskINR",0) > 0 else DARK)
    info_row(ws1, 14, "NTL Suspected",
             "YES — Field inspection required" if result.get("ntlSuspected") else "No",
             HIGH_RED if result.get("ntlSuspected") else DARK)
    info_row(ws1, 15, "Pre-Invoice Flag",
             "HOLD BILLING" if result.get("preInvoiceFlag") else "Clear",
             HIGH_RED if result.get("preInvoiceFlag") else DARK)

    # ── Score breakdown ────────────────────────────────────────────────────────
    ws1.row_dimensions[17].height = 18
    cell_style(ws1, "A17", "SCORE BREAKDOWN", bold=True, size=9, color="888880")

    sc = result.get("scoreComponents", {})
    ws1.row_dimensions[18].height = 20
    cell_style(ws1, "A18", "Isolation Forest",  size=10)
    cell_style(ws1, "B18", sc.get("isolation_forest", 0), bold=True, size=10)
    ws1.row_dimensions[19].height = 20
    cell_style(ws1, "A19", "TFT Deviation",     size=10)
    cell_style(ws1, "B19", sc.get("tft_deviation", 0),    bold=True, size=10)
    ws1.row_dimensions[20].height = 20
    cell_style(ws1, "A20", "Rules Engine",      size=10)
    cell_style(ws1, "B20", sc.get("rules", 0),            bold=True, size=10)
    ws1.row_dimensions[21].height = 22
    cell_style(ws1, "A21", "TOTAL SCORE",       bold=True, size=10)
    cell_style(ws1, "B21", result.get("riskScore", 0),    bold=True, size=14,
               color=sev_colour)

    # ── Explanation ────────────────────────────────────────────────────────────
    ws1.row_dimensions[23].height = 18
    cell_style(ws1, "A23", "EXPLANATION", bold=True, size=9, color="888880")
    ws1.row_dimensions[24].height = 80
    ws1.merge_cells("A24:H24")
    exp = ws1["A24"]
    exp.value     = result.get("explanation","—")
    exp.font      = Font(name="Arial", size=10, color=DARK)
    exp.alignment = Alignment(horizontal="left", vertical="top",
                              wrap_text=True, indent=1)
    exp.fill      = PatternFill("solid", start_color=LIGHT_GREY)

    set_col_widths(ws1, {"A":22,"B":55,"C":12,"D":12,"E":12,"F":12,"G":12,"H":12})

    # ── Sheet 2: Model Details ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Model Details")
    ws2.sheet_view.showGridLines = False

    ws2.row_dimensions[1].height = 36
    ws2.merge_cells("A1:E1")
    t = ws2["A1"]
    t.value     = "Model Outputs"
    t.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill      = PatternFill("solid", start_color=DARK)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    mo = result.get("modelOutputs", {})

    headers = ["Model","Metric","Value","Status","Note"]
    ws2.row_dimensions[2].height = 20
    for i, h in enumerate(headers, 1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color="888880")
        c.fill      = PatternFill("solid", start_color=LIGHT_GREY)
        c.alignment = Alignment(horizontal="left", vertical="center")

    rows = [
        ["Isolation Forest", "Anomaly Flagged",
         str(mo.get("ifIsAnomaly","—")),
         "ANOMALY" if mo.get("ifIsAnomaly") else "NORMAL",
         "Score < threshold → flagged"],
        ["Isolation Forest", "Anomaly Score",
         str(mo.get("ifScore","—")),
         "Lower = more anomalous",
         "Range: -1.0 to 0.0"],
        ["TFT", "Predicted (P50)",
         f"{mo.get('tftPredictedWh',0):,.0f} Wh",
         "Median forecast", "50th percentile"],
        ["TFT", "Lower Bound (P10)",
         f"{mo.get('tftP10',0):,.0f} Wh",
         "Actual below → SuddenDrop", "10th percentile"],
        ["TFT", "Upper Bound (P90)",
         f"{mo.get('tftP90',0):,.0f} Wh",
         "Actual above → SuddenSpike", "90th percentile"],
        ["TFT", "Deviation %",
         f"{mo.get('tftDeviation',0):.2f}%",
         "HIGH" if abs(mo.get("tftDeviation",0)) > 60 else "MEDIUM" if abs(mo.get("tftDeviation",0)) > 30 else "NORMAL",
         "(actual - p50) / p50 × 100"],
        ["TFT", "Sudden Drop",
         str(mo.get("tftSuddenDrop","—")),
         "FLAGGED" if mo.get("tftSuddenDrop") else "—",
         "Actual < P10"],
        ["TFT", "Sudden Spike",
         str(mo.get("tftSuddenSpike","—")),
         "FLAGGED" if mo.get("tftSuddenSpike") else "—",
         "Actual > P90"],
    ]

    for r, row_data in enumerate(rows, 3):
        ws2.row_dimensions[r].height = 20
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", start_color="F9F9F7")

    set_col_widths(ws2, {"A":18,"B":22,"C":18,"D":20,"E":30})

    # ── Sheet 3: Rule Flags ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Rule Flags")
    ws3.sheet_view.showGridLines = False

    ws3.row_dimensions[1].height = 36
    ws3.merge_cells("A1:D1")
    t = ws3["A1"]
    t.value     = "Validation Rule Flags"
    t.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill      = PatternFill("solid", start_color=DARK)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    rf = result.get("ruleFlags", {})
    flag_info = {
        "billingMismatch":      ("Billing Mismatch",      "Billed amount differs from metered by >2%"),
        "rateSumError":         ("Rate Sum Error",         "Rate1+Rate2+Rate3+Rate4 does not equal Total"),
        "zeroConsumption":      ("Zero Consumption",       "Meter recorded zero energy this month"),
        "flatLine":             ("Flat Line",              "Same value recorded for 3+ consecutive months"),
        "powerFactorLow":       ("PF Below 0.85",          "Power factor below minimum IS threshold"),
        "powerFactorDeviation": ("PF Deviation",           "Power factor dropped >8% from historical avg"),
        "ntlSuspected":         ("NTL Suspected",          "Energy drop + PF degradation simultaneously"),
        "bypassSuspected":      ("Bypass Suspected",       "Energy dropped but peak demand stayed high"),
        "tariffBoundaryGaming": ("Tariff Boundary Gaming", "Reading consistently near slab boundary"),
    }

    headers = ["Rule","Description","Fired","Action"]
    ws3.row_dimensions[2].height = 20
    for i, h in enumerate(headers, 1):
        c = ws3.cell(row=2, column=i, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color="888880")
        c.fill      = PatternFill("solid", start_color=LIGHT_GREY)
        c.alignment = Alignment(horizontal="left", vertical="center")

    for r, (key, (label, desc)) in enumerate(flag_info.items(), 3):
        fired = bool(rf.get(key, False))
        ws3.row_dimensions[r].height = 22
        for col, val in enumerate([label, desc,
                                    "YES" if fired else "No",
                                    "FLAG" if fired else "—"], 1):
            cell = ws3.cell(row=r, column=col, value=val)
            if fired:
                cell.font = Font(name="Arial", bold=(col==3), size=10,
                                  color=HIGH_RED if col in [3,4] else DARK)
                cell.fill = PatternFill("solid", start_color="FEF2F1")
            else:
                cell.font = Font(name="Arial", size=10, color="888880")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    set_col_widths(ws3, {"A":22,"B":50,"C":10,"D":12})

    wb.save(output_path)
    print(f"  Excel report saved → {output_path}")


# ── PDF Report ─────────────────────────────────────────────────────────────────
def generate_pdf(result: dict, output_path: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )

    W, H   = A4
    margin = 20 * mm
    col_w  = W - 2 * margin

    sev       = result.get("severity", "LOW")
    sev_color = {
        "HIGH":   colors.HexColor("#C0392B"),
        "MEDIUM": colors.HexColor("#D29922"),
        "LOW":    colors.HexColor("#27AE60"),
    }.get(sev, colors.HexColor("#27AE60"))

    DARK  = colors.HexColor("#1C1B18")
    MUTED = colors.HexColor("#888880")
    LIGHT = colors.HexColor("#F4F3EF")

    def sty(name, **kw):
        return ParagraphStyle(name, **kw)

    meta_sty  = sty("M", fontName="Times-Roman",      fontSize=8,  textColor=DARK, leading=12, spaceAfter=0)
    score_sty = sty("SC", fontName="Times-Bold", fontSize=14, textColor=DARK, leading=20, spaceAfter=2)
    label_sty = sty("L", fontName="Times-Roman",      fontSize=9,  textColor=DARK, leading=13, spaceAfter=0)
    val_sty   = sty("V", fontName="Times-Bold", fontSize=10, textColor=DARK,  leading=14, spaceAfter=0)
    rec_sty   = sty("R", fontName="Times-Roman",      fontSize=10, textColor=DARK,  leading=15, spaceAfter=0)
    rec_b_sty = sty("RB",fontName="Times-Bold", fontSize=10, textColor=DARK,  leading=15, spaceAfter=0)
    foot_sty  = sty("F", fontName="Times-Roman",      fontSize=7,  textColor=DARK, leading=10, spaceAfter=0)

    mo  = result.get("modelOutputs", {})
    sc  = result.get("scoreComponents", {})
    rf  = result.get("ruleFlags", {})
    dev = result.get("deviationPct", 0)

    fired = [k for k, v in rf.items() if v]
    flag_names = {
        "billingMismatch":      "Billing mismatch",
        "rateSumError":         "Rate sum error",
        "zeroConsumption":      "Zero consumption",
        "flatLine":             "Flat line",
        "powerFactorLow":       "Power factor below 0.85",
        "powerFactorDeviation": "Power factor deviation",
        "ntlSuspected":         "NTL suspected",
        "bypassSuspected":      "Bypass suspected",
        "tariffBoundaryGaming": "Tariff boundary gaming",
    }
    fired_labels = [flag_names.get(f, f) for f in fired]

    p10    = mo.get("tftP10", 0)
    p50    = mo.get("tftPredictedWh", 0)
    p90    = mo.get("tftP90", 0)
    actual = result.get("actualWh", 0)
    tft_status = "above upper bound" if actual > p90 else "below lower bound" if actual < p10 else "within expected range"

    rec = result.get("recommendation", "—")
    rec_text = {
        "FIELD_INSPECTION": "Field inspection is recommended. Billing should be held pending physical verification of the meter.",
        "SCHEDULED_REVIEW": "A scheduled review is recommended within the next billing cycle.",
        "MONITOR":          "No immediate action required. Monitor the next billing cycle.",
        "NO_ACTION":        "No action required.",
    }.get(rec, rec)

    story = []

    # ── meta line ──────────────────────────────────────────────────────────────
    story.append(Paragraph(
        f"{result.get('meterId','—')}  ·  {result.get('billingPeriod','—')}  ·  "
        f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        meta_sty))
    story.append(Spacer(1, 3*mm))

    # ── risk score ─────────────────────────────────────────────────────────────
    story.append(Paragraph(
        f"Risk Score: {result.get('riskScore', 0)} / 100  —  {sev}  —  {result.get('anomalyType','—')}",
        score_sty))
    story.append(Spacer(1, 5*mm))

    # ── key metrics as bullet points ───────────────────────────────────────────
    bullet_data = [
        ("Actual consumption",     f"{actual:,.0f} Wh"),
        ("Model forecast (P50)",   f"{p50:,.0f} Wh"),
        ("Forecast range",         f"{p10:,.0f} – {p90:,.0f} Wh  ({tft_status})"),
        ("Deviation from forecast",f"{('+' if dev > 0 else '')}{dev:.1f}%"),
        ("Historical average",     f"{result.get('historicalAvgWh',0):,.0f} Wh  ({result.get('historyMonths',0)} months)"),
        ("IF anomaly score",        f"{mo.get('ifScore',0):.4f}  ({'flagged' if mo.get('ifIsAnomaly') else 'not flagged'})"),

        ("IF contribution",        f"{sc.get('isolation_forest',0):.1f} / 40 pts"),
        ("TFT contribution",       f"{sc.get('tft_deviation',0)} / 30 pts"),
        ("Rules contribution",     f"{sc.get('rules',0)} / 50 pts"),
        ("Rules fired",            ", ".join(fired_labels) if fired_labels else "None"),
        ("NTL suspected",          "Yes" if result.get("ntlSuspected") else "No"),
        ("Pre-invoice flag",       "Hold billing" if result.get("preInvoiceFlag") else "Clear"),
        ("Revenue at risk",        f"Rs {result.get('revenueAtRiskINR',0):,.2f}"),
    ]

    rows = []
    for label, value in bullet_data:
        rows.append([
            Paragraph(f"• {label}", label_sty),
            Paragraph(value, val_sty),
        ])

    t = Table(rows, colWidths=[65*mm, col_w - 65*mm])
    t.setStyle(TableStyle([
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("ROWBACKGROUNDS",(0,0),(-1,-1), [colors.white, LIGHT]),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))

    # ── recommendation paragraph ───────────────────────────────────────────────
    explanation = result.get("explanation", "")
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"<b>Recommendation: {rec}.</b>  {rec_text}  {explanation}",
        rec_sty))

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=margin, rightMargin=margin,
        topMargin=16*mm, bottomMargin=16*mm,
    )
    doc.build(story)
    print(f"  PDF report saved  → {output_path}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Generate Excel and PDF validation reports from BVEngine results'
    )
    parser.add_argument('--result', required=True,
                        help='Path to result JSON file from parse_and_validate.py')
    parser.add_argument('--format', default='both',
                        choices=['excel','pdf','both'],
                        help='Report format (default: both)')
    parser.add_argument('--out', default=None,
                        help='Output filename prefix (default: same as result file)')
    args = parser.parse_args()

    if not os.path.exists(args.result):
        print(f"ERROR: File not found: {args.result}")
        exit(1)

    with open(args.result) as f:
        result = json.load(f)

    prefix = args.out or Path(args.result).stem.replace('_result','')
    ts     = datetime.now().strftime('%Y%m%d_%H%M%S')

    print(f"\nGenerating reports for meter {result.get('meterId','?')} ...")

    if args.format in ('excel','both'):
        xl_path = f"{prefix}_report_{ts}.xlsx"
        generate_excel(result, xl_path)

    if args.format in ('pdf','both'):
        pdf_path = f"{prefix}_report_{ts}.pdf"
        generate_pdf(result, pdf_path)

    print("\nDone!")


if __name__ == '__main__':
    main()